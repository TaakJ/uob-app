from exception import CustomException
from setup import Folder
import logging
from pathlib import Path
from os.path import join
import glob
import shutil
import pandas as pd
import numpy as np
import openpyxl
import chardet
from io import StringIO
import re
import xlrd
import csv

class convert_2_files:

    async def check_source_files(self) -> None:

        logging.info("Check Source files..")

        set_log = []
        for _dir in self.input_dir:

            status_file = "not_found"
            if glob.glob(_dir, recursive=True):
                status_file = "found"

            record = {
                "module": self.module,
                "input_dir": _dir,
                "status_file": status_file,
                "function": "check_source_files",
            }
            set_log.append(record)
            logging.info(f'Source file: "{_dir}", Status: "{status_file}"')

        self._log_setter(set_log)


    async def retrieve_data_from_source_files(self) -> None:

        logging.info("Retrieve Data from Source files..")

        state = "failed"
        for i, record in enumerate(self.logging):
            record.update({"function": "retrieve_data_from_source_files", "state": state})

            _dir = record["input_dir"]
            types = Path(_dir).suffix
            status_file = record["status_file"]
            try:
                if status_file == "found":
                    if [".xlsx", ".xls"].__contains__(types):
                        logging.info(f'Read Excel file: "{_dir}"')
                        _data = self.excel_data_cleaning(i)

                    else:
                        logging.info(f'Read Text file: "{_dir}"')
                        _data = self.text_data_cleaning(i)
                else:
                    continue
                state = "succeed"
                record.update({"data": _data, "state": state})

            except Exception as err:
                record.update({"errors": err})

            if "errors" in record:
                raise CustomException(errors=self.logging)


    def read_text_files(func):
        def wrapper(*args: tuple, **kwargs: dict) -> dict:

            by_lines = iter(func(*args, **kwargs))
            _data = {}

            rows = 0
            while True:
                try:
                    list_by_lines = []
                    for sheets, data in next(by_lines).items():

                        if sheets == "LDS":
                            if rows == 0:
                                ## herder column
                                list_by_lines = " ".join(data).split(" ")
                            else:
                                ## row value
                                for idx, value in enumerate(data):
                                    if idx == 0:
                                        value = re.sub(r"\s+", ",", value).split(",")
                                        list_by_lines.extend(value)
                                    else:
                                        list_by_lines.append(value)

                        elif sheets == "DOC":
                            if rows == 1:
                                ## herder column
                                list_by_lines = " ".join(data).split(" ")
                            elif rows > 1:
                                ## row value
                                for idx, value in enumerate(data):
                                    if idx == 3:
                                        value = re.sub(r"\s+", ",", value).split(",")
                                        list_by_lines.extend(value)
                                    else:
                                        list_by_lines.append(value)

                        elif sheets == "ADM":
                            ## row value
                            list_by_lines = data

                        if list_by_lines != []:
                            if sheets not in _data:
                                _data[sheets] = [list_by_lines]
                            else:
                                _data[sheets].append(list_by_lines)
                        else:
                            continue
                    rows += 1

                except StopIteration:
                    break
            return _data

        return wrapper


    @read_text_files
    def text_data_cleaning(self, i: int) -> any:

        self.logging[i].update({"function": "text_data_cleaning"})

        _dir = self.logging[i]["input_dir"]
        sheets = self.logging[i]["module"]

        files = open(_dir, "rb")
        encoded = chardet.detect(files.read())["encoding"]
        files.seek(0)
        decode_data = StringIO(files.read().decode(encoded))

        for lines in decode_data:
            regex = re.compile(r"\w+.*")
            find_regex = regex.findall(lines)
            if find_regex != []:
                yield {sheets: re.sub(r"\W\s+","||","".join(find_regex).strip()).split("||")}


    def read_excle_files(func):
        def wrapper(*args: tuple, **kwargs: dict) -> dict:

            by_sheets = iter(func(*args, **kwargs))
            _data = {}

            while True:
                try:
                    for sheets, data in next(by_sheets).items():
                        if not all(dup == data[0] for dup in data) and \
                            not data.__contains__("Centralized User Management : User List."):
                            if sheets not in _data:
                                _data[sheets] = [data]
                            else:
                                _data[sheets].append(data)

                except StopIteration:
                    break
            return _data

        return wrapper


    @read_excle_files
    def excel_data_cleaning(self, i: int) -> any:

        self.logging[i].update({"function": "excel_data_cleaning"})

        workbook = xlrd.open_workbook(self.logging[i]["input_dir"])
        sheet_list = [sheet for sheet in workbook.sheet_names() if sheet != "StyleSheet"]

        for sheets in sheet_list:
            cells = workbook.sheet_by_name(sheets)
            for row in range(0, cells.nrows):
                yield {sheets: [cells.cell(row, col).value for col in range(cells.ncols)]}


    def initial_data_types(self, df: pd.DataFrame) -> pd.DataFrame:
        
        self.logging[-1].update({"function": "initial_data_types"})
        
        df = df.astype({"ApplicationCode": object,
                        "AccountOwner": int,
                        "AccountName": object,
                        "AccountType": object,
                        "EntitlementName": object,
                        "SecondEntitlementName": object,
                        "ThirdEntitlementName": object,
                        "AccountStatus": object,
                        "IsPrivileged": object,
                        "AccountDescription": object,
                        "CreateDate": "datetime64[ms]",
                        "LastLogin": "datetime64[ms]",
                        "LastUpdatedDate": "datetime64[ms]",
                        "AdditionalAttribute": object,
                        })
        df[["CreateDate","LastLogin","LastUpdatedDate"]] = df[["CreateDate","LastLogin","LastUpdatedDate"]]\
            .apply(pd.to_datetime, format="%Y%m%d%H%M%S")
        
        if "remark" in df.columns:
            df = df.loc[df["remark"] != "Removed"]
        else:
            df["remark"] = "Inserted"
        
        state = "succeed"
        self.logging[-1].update({"state": state})
                
        return df    

    def validate_data_change(self, df: pd.DataFrame, change_df: pd.DataFrame) -> dict:
        
        def format_record(record):
            return ("{"+"\n".join('{!r}: {!r},'.format(columns, values) \
                for columns, values in record.items())+"}")

        logging.info("Validate Data Change..")
        self.logging[-1].update({"function": "validate_data_change"})

        self.change_rows = {}
        self.remove_rows = []
        if len(df.index) > len(change_df.index):
            self.remove_rows = [idx for idx in list(df.index) if idx not in list(change_df.index)]

        ## reset index.
        union_index = np.union1d(df.index, change_df.index)
        
        # as starter dataframe for compare
        df = df.reindex(index=union_index, columns=df.columns).iloc[:,:-1]
        
        # change data / new data.
        change_df = change_df.reindex(index=union_index, columns=change_df.columns).iloc[:,:-1]
        
        ## compare data 
        df["count"] = pd.DataFrame(np.where(df.ne(change_df), True, False), index=df.index, columns=df.columns)\
            .apply(lambda x: (x == True).sum(), axis=1)
        
        start_rows = 2
        for idx in union_index:
            if idx not in self.remove_rows:
                record = {}
                #[0] => column
                #[1] => value
                for data, change_data in zip(df.items(), change_df.items()):
                    if df.loc[idx, "count"] != 14:
                        ## No_changed rows.
                        if df.loc[idx, "count"] < 1:  # <=1
                            df.at[idx, data[0]] = data[1][idx]
                            df.loc[idx, "remark"] = "No_changed"
                            
                        else:
                            ## Updated rows.
                            if data[1][idx] != change_data[1][idx]:
                                record.update({data[0]: f"{data[1][idx]} => {change_data[1][idx]}"})
                            df.at[idx, data[0]] = change_data[1][idx]
                            df.loc[idx, "remark"] = "Updated"
                            
                    else:
                        ## Inserted rows.
                        record.update({data[0]: change_data[1][idx]})
                        df.at[idx, data[0]] = change_data[1][idx]
                        df.loc[idx, "remark"] = "Inserted"
                        
                if record != {}:
                    self.change_rows[start_rows + idx] = format_record(record)
            else:
                ## Removed rows.
                df.loc[idx, "remark"] = "Removed"
        # self.remove_rows = [idx + start_rows for idx in self.remove_rows]

        df = df.drop(["count"], axis=1)
        df.index += start_rows
        rows_data = df.to_dict(orient='index')

        state = "succeed"
        self.logging[-1].update({"state": state})
        
        return rows_data
    
    
    async def write_data_to_tmp_file(self) -> None:

        logging.info("Write Data to Tmp files..")

        state = "failed"
        for record in self.logging:
            try:
                if record["module"] == "Target_file":

                    tmp_name = join(Folder.TMP,f"TMP_{self.module}-{self.batch_date.strftime('%d%m%y')}.xlsx")
                    record.update({"input_dir": tmp_name,"function": "write_data_to_tmp_file","state": state})

                    data = record["data"]
                    change_df = pd.DataFrame(data)
                    ## check intial data types for new data.
                    change_df = self.initial_data_types(change_df)
                    
                    try:
                        workbook = openpyxl.load_workbook(tmp_name)
                        get_sheet = workbook.get_sheet_names()
                        sheet_num = len(get_sheet)
                        sheet_name = f"RUN_TIME_{sheet_num - 1}"
                        sheet = workbook.get_sheet_by_name(sheet_name)
                        workbook.active = sheet_num

                    except FileNotFoundError:
                        template_name = join(Folder.TEMPLATE, "Application Data Requirements.xlsx")
                        try:
                            if not glob.glob(tmp_name, recursive=True):
                                shutil.copy2(template_name, tmp_name)
                                state = "succeed"
                        except:
                            raise
                        workbook = openpyxl.load_workbook(tmp_name)
                        sheet = workbook.worksheets[0]
                        sheet_name = "RUN_TIME_1"
                        sheet_num = 1
                        sheet.title = sheet_name

                    if state != "succeed":
                        sheet_name = f"RUN_TIME_{sheet_num}"
                        sheet = workbook.create_sheet(sheet_name)
                        
                    # read tmp files.
                    data = sheet.values
                    columns = next(data)[0:]
                    tmp_df = pd.DataFrame(data, columns=columns)
                    ## check intial data types for existing data.
                    tmp_df = self.initial_data_types(tmp_df)
                    
                    logging.info(f'Generate Sheet_name: "{sheet_name}" in Tmp files.')
                    
                    ## validate data change row by row
                    rows_data = self.validate_data_change(tmp_df, change_df)
                    
                    ## write to tmp files.
                    state = self.write_worksheet(sheet, rows_data)
                    workbook.move_sheet(workbook.active, offset=-sheet_num)
                    workbook.save(tmp_name)

                    record.update({"sheet_name": sheet_name, "state": state})
                    logging.info(f"Write to Tmp files state: {state}.")

            except Exception as err:
                record.update({"errors": err})

            if "errors" in record:
                raise CustomException(errors=self.logging)


    def write_worksheet(self, sheet:any, change_data:dict) -> str:

        self.logging[-1].update({"function": "write_worksheet"})
        max_rows = max(change_data, default=0)
        logging.info(f"Data for write: {max_rows}. rows")

        start_rows = 2
        try:
            # write columns.
            for idx, columns in enumerate(change_data[start_rows].keys(), 1):
                sheet.cell(row=1, column=idx).value = columns
                
            ## write data.
            while start_rows <= max_rows:
                for idx, columns in enumerate(change_data[start_rows].keys(),1):
                    if columns == "remark":
                        if start_rows in self.remove_rows and change_data[start_rows][columns] == "Removed":
                            ## Removed data.
                            _show = f"{change_data[start_rows][columns]} Rows: ({start_rows}) in Tmp files."
                            sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                            
                        elif start_rows in self.change_rows.keys() and change_data[start_rows][columns] in ["Inserted","Updated"]:
                            ## Updated / Insert data. 
                            _show = f"{change_data[start_rows][columns]} Rows: ({start_rows}) in Tmp files.\nRecord Changed: {self.change_rows[start_rows]}"
                            sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                        else:
                            ## No change data.
                            _show = f"No Change Rows: ({start_rows}) in Tmp files."
                            sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                            
                        logging.info(_show)
                        
                    elif columns in ["CreateDate","LastUpdatedDate"]:
                        sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns].strftime("%Y%m%d%H%M%S")
                        
                    else:
                        sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                        
                start_rows += 1
            state = "succeed"
            self.logging[-1].update({"state": state})
            
        except KeyError as err:
            raise KeyError(f"Can not Write rows: {err} in Tmp files.")
        
        return state


    async def write_data_to_target_file(self) -> None:

        logging.info("Write Data to Target files..")

        state = "failed"
        for record in self.logging:
            try:
                
                if record["module"] == "Target_file":
                    record.update({"function": "write_data_to_target_file", "state": state})
                    
                    if self.store_tmp is True:
                        tmp_name = record["input_dir"]
                        sheet_name = record["sheet_name"]
                        change_df = pd.read_excel(tmp_name, sheet_name=sheet_name,)
                    else:
                        data = record["data"]
                        change_df = pd.DataFrame(data)
                    
                    ## check intial data types for new data.
                    change_df = self.initial_data_types(change_df)
                    
                    ## set target name for read csv.
                    if self.write_mode == "overwrite" or self.manual:
                        target_name = join(self.output_dir, self.output_file)
                    else:
                        suffix = f"{self.batch_date.strftime('%d%m%y')}"
                        self.output_file = f"{Path(self.output_file).stem}_{suffix}.csv"
                        target_name = join(self.output_dir, self.output_file)
                    
                    ## read csv.    
                    target_df = self.read_csv(target_name)
                    _data = self.optimize_data(target_df, change_df)
                    
                    ## write csv. 
                    state = self.write_csv(target_name, _data)
                    logging.info(f"Write to Target Files status: {state}.")
            
            except Exception as err:
                record.update({"errors": err})

        if "errors" in record:
            raise CustomException(errors=self.logging)
        
        
    def read_csv(self, target_name:str) -> pd.DataFrame:
        
        logging.info(f'Read Target files: "{target_name}"')
        
        state = "failed"
        self.logging[-1].update({"input_dir": target_name, "function": "read_csv", "state": state})
        
        try:
            data = []
            with open(target_name, 'r', newline='') as reader:
                csv_reader = csv.reader(reader, skipinitialspace=True)
                header = next(csv_reader)
                
                for row in csv_reader:
                    data.append(row)
                target_df = pd.DataFrame(data, columns=header)
                                
        except FileNotFoundError:
            template_name = join(Folder.TEMPLATE, "Application Data Requirements.xlsx")
            target_df = pd.read_excel(template_name)
            target_df.to_csv(target_name, index=None, header=True)
        
        state = "succeed"
        self.logging[-1].update({"state": state})
        
        return target_df
        
    
    def optimize_data(self,target_df: pd.DataFrame, change_df: pd.DataFrame) -> dict:

        logging.info("Optimize Data Before Write To Target..")
        
        state = "failed"
        self.logging[-1].update({"function": "optimize_data", "state":state})
        
        _data = {}
        try:
            ## check intial data types for existing data.
            target_df = self.initial_data_types(target_df)
            
            ## filter data on batch date => (DataFrame)
            batch_df = target_df[target_df["CreateDate"].isin(np.array([pd.Timestamp(self.fmt_batch_date)]))]\
                .reset_index(drop=True)
            
            ## filter data not on batch date => (dict).
            _dict = target_df[~target_df["CreateDate"].isin(np.array([pd.Timestamp(self.fmt_batch_date)]))]\
                .iloc[:,:-1].to_dict("index")
            
            ## validate data change row by row
            rows_data = self.validate_data_change(batch_df, change_df)
            
            ## merge data from new and old data.
            max_rows = max(_dict, default=0)
            for idx, values in rows_data.items():
                if idx in self.change_rows or idx in self.remove_rows:
                    values.update({"mark_row": idx})
                _dict = {**_dict, **{max_rows + idx: values}}
            
            ## sorted order data on batch date.
            i = 0
            start_rows = 2
            for idx, values in enumerate(sorted(_dict.values(), key=lambda d: d["CreateDate"])):
                idx += start_rows
                if "mark_row" in values.keys():
                    if values["mark_row"] in self.change_rows:
                        self.change_rows[str(idx)] = self.change_rows.pop(values["mark_row"])
                    elif values["mark_row"] in self.remove_rows:
                        self.remove_rows[i] = idx
                        i += 1
                    values.pop("mark_row")
                _data.update({idx: values})
            
        except Exception as err:
            raise Exception(err)
        
        state = "succeed"
        self.logging[-1].update({"state": state})
        
        return _data
    
    
    def write_csv(self, target_name:str, _data:dict) -> str:
        
        logging.info(f'Write mode: "{self.write_mode}" in Target files: "{target_name}"')
        
        state = "failed"
        self.logging[-1].update({"function": "write_csv", "state": state})
        
        # ## set data types column.
        # df = pd.DataFrame.from_dict(rows_data, orient="index")
        # df[["CreateDate","LastUpdatedDate"]] = df[["CreateDate","LastUpdatedDate"]]\
        #     .apply(lambda d: d.dt.strftime("%Y%m%d%H%M%S"))
        # rows_data = df.to_dict(orient='index')
        try:
            reader = self.read_csv(target_name)
            fieldnames = reader.columns.tolist()

            print(fieldnames)
            # fieldnames = _data[2].keys()
            # print(fieldnames)
            # with open(target_name, 'w', newline='') as writer:
            #     csv_writer = csv.DictWriter(writer, fieldnames=fieldnames)
            #     csv_writer.writeheader()
                
            #     for idx in wt_data:
            #         ## update / insert rows.
            #         if str(idx) in self.change_rows.keys() and wt_data[idx]["remark"] in ["Updated", "Inserted"]:
            #             logging.info(f'"{wt_data[idx]["remark"]}" Rows: "{idx}" in Target file.\nRecord Changed:"{self.change_rows[str(idx)]}"')
            #         else:
            #             continue
                    
            #         if idx not in self.remove_rows:
            #             csv_writer.writerow(wt_data[idx])
                        
            # writer.closed
            # state = "succeed"
                
                
        except Exception as err:
            raise Exception(err)
        
        
        #     ## write csv file.
        #     with open(target_name, 'w', newline='') as writer:
        #         csvout = csv.DictWriter(writer, csvin.fieldnames, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        #         csvout.writeheader()
        #         for idx in rows:            
        #             if idx not in self.remove_rows:
        #                 csvout.writerow(rows[idx])
        #     writer.closed 
        #     state = "succeed"
            
        
        return state


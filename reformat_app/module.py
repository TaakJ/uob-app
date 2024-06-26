import logging
from pathlib import Path
from os.path import join
import glob
import shutil
import pandas as pd
import numpy as np
import openpyxl
import chardet
from io import StringIO
import xlrd
import csv
from .exception import CustomException 
from .setup import Folder

class convert_2_files:

    async def check_source_files(self) -> None:

        logging.info("Check Source files.")

        set_log = []
        for input_dir in self.input_dir:

            status_file = "not_found"
            if glob.glob(input_dir, recursive=True):
                status_file = "found"

            record = {
                "module": self.module,
                "input_dir": input_dir,
                "status_file": status_file,
                "function": "check_source_files",
            }
            set_log.append(record)
            logging.info(f'Source file: "{input_dir}", Status: "{status_file}"')

        self.log_setter(set_log)

    async def retrieve_data_from_source_files(self) -> None:

        logging.info("Retrieve Data from Source files.")

        state = "failed"
        for i, record in enumerate(self.logging):
            record.update({"function": "retrieve_data_from_source_files","state": state})

            input_dir = record["input_dir"]
            types = Path(input_dir).suffix
            status_file = record["status_file"]
            try:
                if status_file == "found":
                    if [".xlsx", ".xls"].__contains__(types):
                        logging.info(f'Read Excel file: "{input_dir}"')
                        data = self.read_excel_file(i)
                    else:
                        logging.info(f'Read Text file: "{input_dir}"')
                        data = self.read_text_file(i)
                else:
                    raise FileNotFoundError(f"status_file: {status_file}")

                state = "succeed"
                record.update({"data": data, "state": state})

            except Exception as err:
                record.update({"err": err})

            if "err" in record:
                raise CustomException(err=self.logging)

    def read_text_file(self, i: int) -> any:

        self.logging[i].update({"function": "read_text_file"})
        input_dir = self.logging[i]["input_dir"]
        module = self.logging[i]["module"]

        file = open(input_dir, "rb")
        encoded = chardet.detect(file.read())["encoding"]
        file.seek(0)
        decode_data = StringIO(file.read().decode(encoded))

        if module == "ADM":
            data = self.extract_adm_data(i, decode_data)
            return data

        elif module == "DOC":
            data = self.extract_doc_data(i, decode_data)
            return data

        elif module == "LDS":
            data = self.extract_lds_data(i, decode_data)
            return data

    def read_excel_file(self, i: int) -> any:

        self.logging[i].update({"function": "read_excel_file"})
        input_dir = self.logging[i]["input_dir"]
        module = self.logging[i]["module"]

        workbook = xlrd.open_workbook(input_dir)

        if module == "BOS":
            data = self.extract_bos_data(i, workbook)
            return data

        elif module == "CUM":
            data = self.extract_cum_data(i, workbook)
            return data

        elif module == "ICA":
            data = self.extract_ica_data(i, workbook)
            return data

        elif module == "IIC":
            data = self.extract_iic_data(i, workbook)
            return data

        elif module == "LMT":
            data = self.extract_lmt_data(i, workbook)
            return data

        elif module == "MOC":
            data = self.extract_moc_data(i, workbook)
            return data

    def initial_data_types(self, df: pd.DataFrame) -> pd.DataFrame:

        state = "failed"
        self.logging[-1].update({"function": "initial_data_types", "state": state})
        try:
            df = df.astype(
                {
                    "ApplicationCode": object,
                    "AccountOwner": object,
                    "AccountName": object,
                    "AccountType": object,
                    "EntitlementName": object,
                    "SecondEntitlementName": object,
                    "ThirdEntitlementName": object,
                    "AccountStatus": object,
                    "IsPrivileged": object,
                    "AccountDescription": object,
                    "CreateDate": "datetime64[ms]",
                    "LastLogin": "datetime64[ms]",
                    "LastUpdatedDate": "datetime64[ms]",
                    "AdditionalAttribute": object,
                }
            )
            df[["CreateDate", "LastLogin", "LastUpdatedDate"]] = df[["CreateDate", "LastLogin", "LastUpdatedDate"]]\
                .apply(pd.to_datetime, format="%Y%m%d%H%M%S")

            if "remark" in df.columns:
                df = df.loc[df["remark"] != "Remove"]
            else:
                df["remark"] = "Insert"

        except Exception as err:
            raise Exception(err)

        state = "succeed"
        self.logging[-1].update({"state": state})

        return df

    def validate_data_change(self, df: pd.DataFrame, change_df: pd.DataFrame) -> dict:

        logging.info("Validate Data Change..")

        self.change_rows = {}
        self.remove_rows = []
        state = "failed"
        self.logging[-1].update({"function": "validate_data_change", "state": state})

        def format_record(record):
            return (
                "{"
                + "\n".join(
                    "{!r} => {!r},".format(columns, values) for columns, values in record.items()
                )
                + "}"
            )

        if len(df.index) > len(change_df.index):
            self.remove_rows = [idx for idx in list(df.index) if idx not in list(change_df.index)]

        try:
            ## merge index.
            merge_index = np.union1d(df.index, change_df.index)
            ## as starter dataframe for compare.
            df = df.reindex(index=merge_index, columns=df.columns).iloc[:, :-1]
            ## change data / new data.
            change_df = change_df.reindex(index=merge_index, columns=change_df.columns).iloc[:, :-1]

            ## compare data.
            df["count"] = pd.DataFrame(np.where(df.ne(change_df), True, df),index=df.index,columns=df.columns)\
                .apply(lambda x: (x == True).sum(), axis=1)

            i = 0
            start_rows = 2
            for idx in merge_index:
                if idx not in self.remove_rows:
                    record = {}
                    ## data[0] => column
                    ## data[1] => value
                    for data, change_data in zip(df.items(), change_df.items()):
                        if df.loc[idx, "count"] != 14:
                            if df.loc[idx, "count"] < 1:
                                df.loc[idx, data[0]] = data[1][idx]
                                df.loc[idx, "remark"] = "No_change"

                            else:
                                if data[1][idx] != change_data[1][idx]:
                                    record.update({data[0]: change_data[1][idx]})
                                df.loc[idx, data[0]] = change_data[1][idx]
                                df.loc[idx, "remark"] = "Update"

                        else:
                            record.update({data[0]: change_data[1][idx]})
                            df.loc[idx, data[0]] = change_data[1][idx]
                            df.loc[idx, "remark"] = "Insert"

                    if record != {}:
                        self.change_rows[start_rows + idx] = format_record(record)

                else:
                    self.remove_rows[i] = start_rows + idx
                    df.loc[idx, "remark"] = "Remove"
                    i += 1

            ## set dataframe.
            df = df.drop(["count"], axis=1)
            df.index += start_rows
            data_dict = df.to_dict(orient="index")

        except Exception as err:
            raise Exception(err)

        state = "succeed"
        self.logging[-1].update({"state": state})

        return data_dict

    async def write_data_to_tmp_file(self) -> None:

        logging.info("Write Data to Tmp files..")

        state = "failed"
        for record in self.logging:
            try:
                if record["module"] == "Target_file":

                    tmp_name = join(Folder.TMP,f"TMP_{self.module}-{self.batch_date.strftime('%Y%m%d')}.xlsx")
                    record.update({"input_dir": tmp_name, "function": "write_data_to_tmp_file","state": state})

                    data = record["data"]
                    change_df = pd.DataFrame(data)
                    change_df = self.initial_data_types(change_df)

                    try:
                        workbook = openpyxl.load_workbook(tmp_name)
                        get_sheet = workbook.get_sheet_names()
                        sheet_num = len(get_sheet)
                        sheet_name = f"RUN_TIME_{sheet_num - 1}"
                        sheet = workbook.get_sheet_by_name(sheet_name)
                        workbook.active = sheet_num

                    except FileNotFoundError:
                        template_name = join(Folder.TEMPLATE,"Application Data Requirements.xlsx")
                        try:
                            if not glob.glob(tmp_name, recursive=True):
                                shutil.copy2(template_name, tmp_name)
                                state = "succeed"
                        except:
                            raise
                        workbook = openpyxl.load_workbook(tmp_name)
                        sheet = workbook.worksheets[0]
                        sheet_name = "RUN_TIME_1"
                        sheet_num = 1
                        sheet.title = sheet_name

                    ## read tmp files.
                    data = sheet.values
                    columns = next(data)[0:]
                    tmp_df = pd.DataFrame(data, columns=columns)
                    tmp_df = self.initial_data_types(tmp_df)

                    ## validate data change row by row
                    data_dict = self.validate_data_change(tmp_df, change_df)

                    ## write to tmp files.
                    if state != "succeed":
                        sheet_name = f"RUN_TIME_{sheet_num}"
                        sheet = workbook.create_sheet(sheet_name)
                    logging.info(f'Generate Sheet_name: "{sheet_name}" in Tmp files.')

                    state = self.write_worksheet(sheet, data_dict)
                    workbook.move_sheet(workbook.active, offset=-sheet_num)
                    workbook.save(tmp_name)

                    record.update({"sheet_name": sheet_name, "state": state})
                    logging.info(f'Write Data to Tmp files: "{state}".')

            except Exception as err:
                record.update({"err": err})

            if "err" in record:
                raise CustomException(err=self.logging)

    def write_worksheet(self, sheet: any, change_data: dict) -> str:

        logging.info("Write to Worksheet.")

        state = "failed"
        self.logging[-1].update({"function": "write_worksheet", "state": state})

        start_rows = 2
        max_rows = max(change_data, default=0)
        try:
            # write columns.
            for idx, columns in enumerate(change_data[start_rows].keys(), 1):
                sheet.cell(row=1, column=idx).value = columns

            ## write data.
            while start_rows <= max_rows:
                for idx, columns in enumerate(change_data[start_rows].keys(), 1):

                    if columns == "remark":
                        if (start_rows in self.remove_rows and change_data[start_rows][columns] == "Remove"):
                            ## Remove rows.
                            show = f'{change_data[start_rows][columns]} Rows: "{start_rows}" in Tmp files.'
                            sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                        elif start_rows in self.change_rows.keys() and change_data[start_rows][columns] in ["Insert", "Update"]:
                            ## Update / Insert rows.
                            show = f'{change_data[start_rows][columns]} Rows: "{start_rows}" in Tmp files.\nRecord Change: {self.change_rows[start_rows]}'
                            sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                        else:
                            ## No change rows.
                            show = f'No Change Rows: "{start_rows}" in Tmp files.'
                            sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]

                        logging.info(show)

                    else:
                        sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]

                start_rows += 1

        except KeyError as err:
            raise KeyError(f"Can not Write rows: {err} in Tmp files.")

        state = "succeed"
        self.logging[-1].update({"state": state})

        return state

    async def write_data_to_target_file(self) -> None:

        logging.info("Write Data to Target files..")
        state = "failed"
        for record in self.logging:
            try:

                if record["module"] == "Target_file":
                    record.update({"function": "write_data_to_target_file","state": state})

                    if self.store_tmp is True:
                        tmp_name = record["input_dir"]
                        sheet_name = record["sheet_name"]
                        change_df = pd.read_excel(tmp_name,sheet_name=sheet_name,dtype=object)
                    else:
                        data = record["data"]
                        change_df = pd.DataFrame(data)

                    ## check intial data types for new data.
                    change_df = self.initial_data_types(change_df)

                    ## set target name for read csv.
                    if self.write_mode == "overwrite" or self.manual:
                        target_name = join(self.output_dir, self.output_file)
                    else:
                        suffix = f"{self.batch_date.strftime('%Y%m%d')}"
                        self.output_file = f"{Path(self.output_file).stem}_{suffix}.csv"
                        target_name = join(self.output_dir, self.output_file)

                    ## read csv.
                    target_df = self.read_csv(target_name)
                    output = self.optimize_data(target_df, change_df)

                    ## write csv.
                    state = self.write_csv(target_name, output)
                    logging.info(f'Write to Target Files status: "{state}"')

            except Exception as err:
                record.update({"err": err})

        if "err" in record:
            raise CustomException(err=self.logging)

    def read_csv(self, target_name: str) -> pd.DataFrame:

        logging.info(f'Read Target files: "{target_name}"')

        state = "failed"
        self.logging[-1].update({"input_dir": target_name,"function": "read_csv","state": state})

        try:
            data = []
            with open(target_name, "r", newline="") as reader:
                csv_reader = csv.reader(reader,
                                        skipinitialspace=True,
                                        quoting=csv.QUOTE_ALL,
                                        quotechar='"')
                header = next(csv_reader)

                for row in csv_reader:
                    data.append(row)
                target_df = pd.DataFrame(data, columns=header)

        except FileNotFoundError:
            template_name = join(Folder.TEMPLATE, "Application Data Requirements.xlsx")
            target_df = pd.read_excel(template_name)
            target_df.to_csv(target_name, index=None, header=True)

        state = "succeed"
        self.logging[-1].update({"state": state})

        return target_df

    def optimize_data(self, target_df: pd.DataFrame, change_df: pd.DataFrame) -> dict:

        logging.info("Optimize Data Before Write to Target..")

        state = "failed"
        self.logging[-1].update({"function": "optimize_data", "state": state})

        output = {}
        try:
            ## check intial data types for existing data.
            target_df = self.initial_data_types(target_df)

            ## filter data on batch date => (DataFrame).
            batch_df = target_df[target_df["CreateDate"]\
                .isin(np.array([pd.Timestamp(self.fmt_batch_date)]))]\
                .reset_index(drop=True)

            ## filter data not on batch date => (dict).
            merge_data = target_df[~target_df["CreateDate"]\
                .isin(np.array([pd.Timestamp(self.fmt_batch_date)]))]\
                .iloc[:, :-1].to_dict("index")

            ## validate data change row by row.
            data_dict = self.validate_data_change(batch_df, change_df)

            ## merge data from new and old data.
            max_rows = max(merge_data, default=0)
            for idx, values in data_dict.items():
                if idx in self.change_rows or idx in self.remove_rows:
                    values.update({"mark_row": idx})
                merge_data = {**merge_data, **{max_rows + idx: values}}

            ## sorted order data on batch date.
            i = 0
            start_rows = 2
            for idx, values in enumerate(sorted(merge_data.values(), key=lambda d: d["CreateDate"])):
                idx += start_rows
                if "mark_row" in values.keys():
                    if values["mark_row"] in self.change_rows:
                        self.change_rows[str(idx)] = self.change_rows.pop(values["mark_row"])

                    elif values["mark_row"] in self.remove_rows:
                        self.remove_rows[i] = idx
                        i += 1
                    values.pop("mark_row")

                output.update({idx: values})

        except Exception as err:
            raise Exception(err)

        state = "succeed"
        self.logging[-1].update({"state": state})

        return output

    def write_csv(self, target_name: str, output: dict) -> str:

        logging.info(f'Write mode: "{self.write_mode}" in Target files: "{target_name}"')

        state = "failed"
        self.logging[-1].update({"function": "write_csv", "state": state})

        try:
            start_row = 2
            with open(target_name, "r", newline="") as reader:
                csvin = csv.DictReader(reader,
                                    skipinitialspace=True,
                                    quoting=csv.QUOTE_ALL,
                                    quotechar='"')
                rows = {idx + start_row: value for idx, value in enumerate(csvin)}
                for idx in output:
                    data = {columns: values for columns, values in output[idx].items() if columns != "remark"}

                    if str(idx) in self.change_rows.keys() and output[idx]["remark"] in ["Update","Insert"]:
                        logging.info(f'"{output[idx]["remark"]}" Rows: "{idx}" in Target file.\nRecord Change:"{self.change_rows[str(idx)]}"')
                        rows.update({idx: data})
                    else:
                        if idx in self.remove_rows:
                            continue
                        rows[idx].update(data)

            # write csv file.
            with open(target_name, "w", newline="") as writer:
                csvout = csv.DictWriter(writer,
                                        csvin.fieldnames,
                                        quoting=csv.QUOTE_ALL,
                                        quotechar='"')
                csvout.writeheader()
                for idx in rows:
                    if idx not in self.remove_rows:
                        rows[idx].update({
                                "CreateDate": rows[idx]["CreateDate"].strftime("%Y%m%d%H%M%S"),
                                "LastLogin": rows[idx]["LastLogin"].strftime("%Y%m%d%H%M%S"),
                                "LastUpdatedDate": rows[idx]["LastUpdatedDate"].strftime("%Y%m%d%H%M%S"),
                            })
                        csvout.writerow(rows[idx])
            writer.closed

        except Exception as err:
            raise Exception(err)

        state = "succeed"
        self.logging[-1].update({"state": state})

        return state
